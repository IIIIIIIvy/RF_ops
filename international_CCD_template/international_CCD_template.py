import pandas as pd
from tkinter import messagebox
import tkinter as tk
import os
import math
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles import Side, Border

in_ten_dict = {1: 'one', 2: 'two', 3: 'three', 4: 'four', 5: 'five', 6: 'six', 7: 'seven', 8: 'eight', 9: 'nine'}
in_hundred_dict = {2: "twenty", 3: "thirty", 4: "forty", 5: "fifty", 6: "sixty", 7: "seventy", 8: "eighty", 9: "ninety"}
in_twenty_dict = {10: "ten", 11: "eleven", 12: "twelve", 13: "thirteen", 14: "fourteen", 15: "fifteen", 16: "sixteen",
                  17: "seventeen", 18: "eighteen", 19: "nineteen"}


def spell_number_in_thousand(number):
    res_str = ""
    hundred_num = int(number / 100)
    if hundred_num != 0:
        hundred_num_str = in_ten_dict[hundred_num]
        res_str = res_str + hundred_num_str + ' ' + 'hundred'

    number = number % 100
    ten_num = int(number / 10)
    if ten_num == 1:
        res_str = res_str + ' ' + in_twenty_dict[number]
        return res_str
    elif ten_num != 0:
        res_str = res_str + ' ' + in_hundred_dict[ten_num]

    one_num = int(number % 10)
    if one_num != 0:
        res_str = res_str + ' ' + in_ten_dict[one_num]

    return res_str


def spell_number(number):
    res = ""

    num_float = round(math.modf(number)[0], 2) * 100
    num_int = math.modf(number)[1]

    billion_number = int(num_int / 1000000000)
    billion_str = spell_number_in_thousand(billion_number)
    if len(billion_str) != 0:
        res = res + billion_str + " billion"

    num_int = num_int % 1000000000
    million_number = int(num_int / 1000000)
    million_str = spell_number_in_thousand(million_number)
    if len(million_str) != 0:
        res = res + ' ' + million_str + " million"

    num_int = num_int % 1000000
    thousand_number = int(num_int / 1000)
    thousand_str = spell_number_in_thousand(thousand_number)
    if len(thousand_str) != 0:
        res = res + ' ' + thousand_str + " thousand"

    num_int = int(num_int % 1000)
    res = res + ' ' + spell_number_in_thousand(num_int)

    if len(res) == 0:
        res = 'zero'

    res_list = res.split(' ')
    res_list = list(filter(lambda x: len(x) != 0, res_list))
    res = " ".join(res_list)

    if not number.is_integer():
        res = res + ' AND ' + str(int(num_float)) + '/100 '
    return res.upper()


def data_extraction(file_dir, data_file_name):
    NON_EDI = pd.read_excel(file_dir + data_file_name, sheet_name='Non EDI', dtype=str)
    NON_EDI = NON_EDI[['Order Number', 'Related Order Number', 'Sold To Name', 'Customer PO',
                       '2nd Item Number', 'Quantity', 'First Ship Date', 'Ship To', 'Unit Price', 'Sold To',
                       'Supplier Name']]
    NON_EDI.rename(columns={'First Ship Date': 'Date'}, inplace=True)
    NON_EDI['Invoice No'] = NON_EDI['Order Number']
    NON_EDI.fillna('', inplace=True)
    # NON_EDI

    sold_to = pd.read_excel(file_dir + data_file_name, sheet_name='Sold to', dtype=str)
    sold_to.fillna('', inplace=True)
    sold_to['ADDRESS Line 5'] = sold_to['City'] + ',' + sold_to['State'] + ',' + sold_to['Postal Code'] + ',' + sold_to[
        'Country']
    sold_to['ADDRESS Line 5'] = sold_to['ADDRESS Line 5'].apply(
        lambda x: ','.join(list(filter(lambda s: len(s) != 0, x.split(',')))))
    sold_to['sold_to address info'] = sold_to['Alpha Name '] + ";" + sold_to['ADDRESS Line 1'] + ";" + sold_to[
        'ADDRESS Line 2'] + ";" \
                                      + sold_to['ADDRESS Line 3'] + ";" + sold_to['ADDRESS Line 4'] + ";" + sold_to[
                                          'ADDRESS Line 5']
    sold_to['discount info'] = sold_to.apply(lambda x: sorted([x['DISCOUNT 1'], x['DISCOUNT 2']], reverse=True), axis=1)

    ship_to = pd.read_excel(file_dir + data_file_name, sheet_name='ship to', dtype=str)
    ship_to.fillna('', inplace=True)
    ship_to['ADDRESS Line 5'] = ship_to['City'] + ',' + ship_to['State'] + ',' + ship_to['Postal Code'] + ',' + ship_to[
        'Country']
    ship_to['ADDRESS Line 5'] = ship_to['ADDRESS Line 5'].apply(
        lambda x: ','.join(list(filter(lambda s: len(s) != 0, x.split(',')))))
    ship_to['ship_to address info'] = ship_to['Alpha Name '] + ";" + ship_to['ADDRESS Line 1'] + ";" + ship_to[
        'ADDRESS Line 2'] + ";" \
                                      + ship_to['ADDRESS Line 3'] + ";" + ship_to['ADDRESS Line 4'] + ";" + ship_to[
                                          'ADDRESS Line 5']
    del ship_to['Search Type']

    vendor = pd.read_excel(file_dir + data_file_name, sheet_name='Vendor', dtype=str)
    vendor.fillna('', inplace=True)
    vendor['vendor address info'] = vendor['Factory Name'] + ";" + vendor['Address 1'] + ";" + vendor['Address 2'] + ";" \
                                    + vendor['Address 3'] + ";" + vendor['Address 4']

    item_master = pd.read_excel(file_dir + data_file_name, sheet_name='Item master', header=7, usecols='A:BF',
                                dtype=str)
    item_master.fillna('', inplace=True)

    item_master['sku list'] = item_master['SKU List \n(LG Software)'].apply(lambda x: x.split(','))
    item_master['sku list len'] = item_master['SKU List \n(LG Software)'].apply(lambda x: len(x.split(',')))
    single_SKU_df = item_master[item_master['sku list len'] == 1]

    mul_SKU_data = item_master[item_master['sku list len'] != 1]
    mul_SKU_df = pd.DataFrame()
    for index, rows in mul_SKU_data.iterrows():
        data = pd.DataFrame(rows)
        res = pd.concat([data.T for i in range(rows['sku list len'])]).reset_index(drop=True)
        for i in range(rows['sku list len']):
            res.loc[i, 'SKU List \n(LG Software)'] = rows['sku list'][i]
        mul_SKU_df = pd.concat([mul_SKU_df, res])
    mul_SKU_df['SKU List \n(LG Software)']=mul_SKU_df['SKU List \n(LG Software)'].apply(lambda x:x.strip())

    item_master = pd.concat([single_SKU_df, mul_SKU_df]).reset_index(drop=True)
    item_master['SKU List \n(LG Software)'] = item_master['SKU List \n(LG Software)'].apply(
        lambda x: x.replace(' ', ''))

    common_name = pd.read_excel(file_dir + data_file_name, sheet_name='Common Name ', usecols='A:G', dtype=str)
    common_name.fillna('', inplace=True)
    common_name = common_name[['Address Number', 'Second Item Number', 'Cross Reference Item Number']]

    data = pd.merge(NON_EDI, sold_to[['SOLD TO#', 'sold_to address info', 'PAYMENT TERM', 'discount info']],
                    left_on='Sold To', right_on='SOLD TO#', how='left')
    del data['SOLD TO#']

    data = pd.merge(data, ship_to[['SHIP TO #', 'ship_to address info']], left_on='Ship To', right_on='SHIP TO #',
                    how='left')
    del data['SHIP TO #']

    data = pd.merge(data, vendor[['Factory ID', 'Factory Name', 'vendor address info',
                                  'FOB PORT', 'Origin Country', 'JDE SUPPLIER']], left_on='Supplier Name',
                    right_on='JDE SUPPLIER', how='left')
    data['FOB'] = data['FOB PORT'] + "," + data['Origin Country']
    data.rename(columns={'Origin Country': 'COUNTRY OF ORIGIN'}, inplace=True)
    del data['FOB PORT']

    data = pd.merge(data, item_master, left_on=['Factory ID', '2nd Item Number'],
                    right_on=['Factory', 'SKU List \n(LG Software)'], how='left')
    data = pd.merge(data, common_name, left_on=['Sold To', '2nd Item Number'],
                    right_on=['Address Number', 'Second Item Number'], how='left')
    del data['Address Number'], data['Second Item Number']

    data.fillna('', inplace=True)
    data = data.drop_duplicates(['Order Number', '2nd Item Number']).reset_index(drop=True)

    data = data[['Order Number', 'Invoice No', 'Date', 'PAYMENT TERM', 'Customer PO',
                 'sold_to address info', 'ship_to address info',
                 'COUNTRY OF ORIGIN', 'FOB',
                 'discount info',
                 'Factory Name', 'vendor address info',
                 'Quantity', 'Unit Price',
                 '2nd Item Number', 'Model #', 'SKU List \n(LG Software)', 'Description', 'For US', '12 digits UPC',
                 'Cross Reference Item Number',
                 'Qty/\nCarton', 'Net Weight (kg)', 'Gross Weight (kg)', 'Cubic\nMeters (per carton)',
                 'Gross Weight (lbs)', 'Cubic\nFeet (per Carton)',
                 'Width (L) cm', 'Depth (W) cm', 'Height (H) cm',
                 'sku list', 'sku list len']]

    for i in data['Order Number'].unique():
        temp = data[data['Order Number'] == i]
        for j in temp['Factory Name'].unique():
            temp2 = temp[temp['Factory Name'] == j]
            model_list = temp2[temp2['2nd Item Number'] != 'MOQ'].drop_duplicates('2nd Item Number')[
                '2nd Item Number'].to_list()
            data.loc[
                (data['Order Number'] == i) & (data['Factory Name'] == j), 'model_list'] = 'MODLE#: ' + ', '.join(
                model_list)

    data.loc[(data['model_list'] != 'MODLE#: '), 'manufacturer'] = data['vendor address info'] + ';' + data[
        'model_list']

    return data


def data_integration(data):
    table_head_df = data[['Order Number', 'Invoice No', 'Date', 'PAYMENT TERM', 'Customer PO',
                          'sold_to address info', 'ship_to address info', 'COUNTRY OF ORIGIN', 'FOB',
                          'discount info']]
    table_head_dict = table_head_df.drop_duplicates('Order Number').set_index('Order Number').to_dict('index')

    for i in data['Order Number'].unique():
        temp = data[data['Order Number'] == i]
        manufacturer_list = list(temp['manufacturer'].dropna().unique())
        table_head_dict[i].update({'manufacturer_list': manufacturer_list})

    table_content_df = data[
        ['Order Number', '2nd Item Number', 'Quantity', 'Unit Price', 'Description', 'For US', '12 digits UPC',
         'Cross Reference Item Number', 'Qty/\nCarton', 'Net Weight (kg)',
         'Gross Weight (kg)', 'Cubic\nMeters (per carton)', 'Gross Weight (lbs)',
         'Cubic\nFeet (per Carton)', 'Width (L) cm', 'Depth (W) cm',
         'Height (H) cm']]
    table_content_df = table_content_df.drop_duplicates()
    model_dict_list = []
    for i in table_content_df['Order Number'].unique():
        temp = table_content_df[table_content_df['Order Number'] == i]
        model_dict_list.append(
            temp[['2nd Item Number', 'Description', 'For US', '12 digits UPC', 'Cross Reference Item Number',
                  'Quantity', 'Unit Price',
                  'Qty/\nCarton', 'Net Weight (kg)', 'Gross Weight (kg)', 'Cubic\nMeters (per carton)',
                  'Gross Weight (lbs)',
                  'Cubic\nFeet (per Carton)', 'Width (L) cm', 'Depth (W) cm',
                  'Height (H) cm']].to_dict('records'))
        # print(model_dict_list,len(model_dict_list))
    model_dict = dict(zip(table_content_df['Order Number'].unique(), model_dict_list))
    return table_head_dict, model_dict


def write_inv_template(worksheet_inv, temp_dict, temp_content_dict):
    # 定义格式（字体、对齐方式）
    font_title = Font(name="Arail", size=10, bold=True, italic=True)
    font_content = Font(name="Arail", size=10)
    align_title_left = Alignment(horizontal='center', vertical='bottom')
    align_content_left = Alignment(vertical='bottom')
    align_title_right = Alignment(horizontal='right', vertical='bottom')
    align_content_right = Alignment(horizontal='left', vertical='bottom')
    title_row_number = 1

    # ---sold to
    # 单个单元格的值写入
    worksheet_inv['A1'] = 'Sold To'
    worksheet_inv['A1'].font = font_title
    worksheet_inv['A1'].alignment = align_title_left

    # 一列单元格的值写入
    sold_to_address_list = temp_dict['sold_to address info'].split(';')
    writer_pointer = 1
    list_pointer = 0
    for list_pointer in range(len(sold_to_address_list)):
        info = sold_to_address_list[list_pointer]
        if info != '':
            worksheet_inv['B' + str(writer_pointer)] = info
            worksheet_inv['B' + str(writer_pointer)].font = font_content
            worksheet_inv['B' + str(writer_pointer)].alignment = align_content_left

            writer_pointer = writer_pointer + 1
            title_row_number = title_row_number + 1

    # ---ship to
    writer_pointer = writer_pointer + 1
    title_row_number = title_row_number + 1
    # 单个单元格的值写入
    worksheet_inv['A' + str(writer_pointer)] = 'Ship To:'
    worksheet_inv['A' + str(writer_pointer)].font = font_title
    worksheet_inv['A' + str(writer_pointer)].alignment = align_title_left

    # 一列单元格的值写入
    ship_to_address_list = temp_dict['ship_to address info'].split(';')
    list_pointer = 0
    for list_pointer in range(len(ship_to_address_list)):
        info = ship_to_address_list[list_pointer]
        if info != '':
            worksheet_inv['B' + str(writer_pointer)] = info
            worksheet_inv['B' + str(writer_pointer)].font = font_content
            worksheet_inv['B' + str(writer_pointer)].alignment = align_content_left

            writer_pointer = writer_pointer + 1
            title_row_number = title_row_number + 1

    # ---shipped via
    writer_pointer = writer_pointer + 1
    title_row_number = title_row_number + 1
    # 单个单元格的值写入
    worksheet_inv['A' + str(writer_pointer)] = 'Shipped Via:'
    worksheet_inv['A' + str(writer_pointer)].font = font_title
    worksheet_inv['A' + str(writer_pointer)].alignment = align_title_left

    worksheet_inv['B' + str(writer_pointer)] = 'A vessel'
    worksheet_inv['B' + str(writer_pointer)].font = Font(name="Arail", size=10, bold=True)
    worksheet_inv['B' + str(writer_pointer)].alignment = align_content_left

    # ---表头右半部分
    worksheet_inv['G1'] = 'Invoice No:'
    worksheet_inv['G2'] = 'Date:'
    worksheet_inv['G3'] = 'Payment Terms:'
    worksheet_inv['G4'] = 'Customer PO#:'
    worksheet_inv['G6'] = 'COUNTRY OF ORIGIN:'
    worksheet_inv['G7'] = 'FOB:'
    worksheet_inv['H1'] = temp_dict['Invoice No']
    worksheet_inv['H2'] = temp_dict['Date'][:-9]
    worksheet_inv['H2'].number_format = 'yyyy-mm-dd'
    worksheet_inv['H3'] = temp_dict['PAYMENT TERM']
    worksheet_inv['H4'] = temp_dict['Customer PO']
    worksheet_inv['H6'] = temp_dict['COUNTRY OF ORIGIN']
    worksheet_inv['H7'] = temp_dict['FOB']

    for i in range(1, 8):
        worksheet_inv['G' + str(i)].font = font_title
        worksheet_inv['H' + str(i)].font = font_content
        worksheet_inv['G' + str(i)].alignment = align_title_right
        worksheet_inv['H' + str(i)].alignment = align_content_right

    for i in range(1, writer_pointer + 1):
        worksheet_inv.row_dimensions[i].height = 15
    worksheet_inv.row_dimensions[writer_pointer + 1].height = 12

    # ---表格部分！！！
    font_head_title = Font(name="Arail", size=11, bold=True)
    font_important_content = Font(name="Arail", size=10, bold=True)
    font_trivial_content = Font(name="Arail", size=10)

    # 表格表头
    writer_pointer = writer_pointer + 2
    title_row_number = title_row_number + 1
    worksheet_inv['A' + str(writer_pointer)] = 'Marks & Nos.'
    worksheet_inv['B' + str(writer_pointer)] = 'Description'
    worksheet_inv.merge_cells("B" + str(writer_pointer) + ':E' + str(writer_pointer))
    worksheet_inv['F' + str(writer_pointer)] = 'Quantity'
    worksheet_inv['G' + str(writer_pointer)] = 'Unit Price'
    worksheet_inv['H' + str(writer_pointer)] = 'Amount'
    for cell in worksheet_inv[writer_pointer]:
        cell.font = font_head_title
        cell.alignment = Alignment(horizontal='center', vertical='center')
    worksheet_inv.row_dimensions[writer_pointer].height = 16.5

    # 设置线条的样式和颜色
    side = Side(style="medium")
    # 设置单元格的边框线条
    worksheet_inv['A' + str(writer_pointer)].border = Border(bottom=side, top=side, right=side)
    worksheet_inv['B' + str(writer_pointer)].border = Border(bottom=side, top=side)
    worksheet_inv['C' + str(writer_pointer)].border = Border(bottom=side, top=side)
    worksheet_inv['D' + str(writer_pointer)].border = Border(bottom=side, top=side)
    worksheet_inv['E' + str(writer_pointer)].border = Border(bottom=side, top=side, right=side)
    worksheet_inv['F' + str(writer_pointer)].border = Border(bottom=side, top=side, right=side)
    worksheet_inv['G' + str(writer_pointer)].border = Border(bottom=side, top=side, right=side)
    worksheet_inv['H' + str(writer_pointer)].border = Border(bottom=side, top=side)

    # 表格内容
    model_list = temp_content_dict
    model_start_pointer = writer_pointer + 2
    title_row_number = title_row_number + 2

    total_amount = 0
    item_count = 0
    for info in model_list:
        if info['2nd Item Number'] != 'MOQ':
            item_count += 1
            if item_count == 8:
                worksheet_inv.print_title_rows = '1:' + str(title_row_number)
                writer_pointer = 60
            else:
                writer_pointer = writer_pointer + 2

            i = writer_pointer

            worksheet_inv['B' + str(writer_pointer)] = 'Item No.'
            worksheet_inv['C' + str(writer_pointer)] = info['2nd Item Number']
            worksheet_inv['D' + str(writer_pointer)] = '-'
            worksheet_inv['E' + str(writer_pointer)] = info['Description']
            worksheet_inv['F' + str(writer_pointer)] = int(info['Quantity'])
            worksheet_inv['G' + str(writer_pointer)] = float(info['Unit Price'])
            worksheet_inv['H' + str(writer_pointer)] = '=F' + str(writer_pointer) + '*G' + str(writer_pointer)
            worksheet_inv['G' + str(writer_pointer)].number_format = '"$"#,##0.00_-'
            worksheet_inv['H' + str(writer_pointer)].number_format = '"$"#,##0.00_-'
            total_amount = total_amount + int(info['Quantity']) * float(info['Unit Price'])

            writer_pointer = writer_pointer + 1
            worksheet_inv['B' + str(writer_pointer)] = 'HTS NO.:'
            worksheet_inv['C' + str(writer_pointer)] = info['For US']
            writer_pointer = writer_pointer + 1
            worksheet_inv['B' + str(writer_pointer)] = 'UPC NO.:'
            worksheet_inv['C' + str(writer_pointer)] = info['12 digits UPC']
            writer_pointer = writer_pointer + 1
            worksheet_inv['B' + str(writer_pointer)] = 'Customer SKU:'
            worksheet_inv['C' + str(writer_pointer)] = info['Cross Reference Item Number']

            for char in range(ord('B'), ord('I')):
                worksheet_inv[chr(char) + str(i)].font = font_important_content

            for char in range(ord('B'), ord('D')):
                for j in range(i + 1, i + 4):
                    worksheet_inv[chr(char) + str(j)].font = font_trivial_content

    model_end_pointer = writer_pointer
    writer_pointer = writer_pointer + 2
    discount_list = temp_dict['discount info']
    for discount in discount_list:
        if discount != '':
            worksheet_inv['E' + str(writer_pointer)] = discount
            worksheet_inv['H' + str(writer_pointer)] = '=-SUM(H' + str(model_start_pointer) + ':H' + str(
                model_end_pointer) + ')*(LEFT(E' + str(writer_pointer) + ',FIND("%",E' + str(
                writer_pointer) + ')-1)/100)'
            total_amount = total_amount * (1 - float(discount[:discount.find('%')]) / 100)

            worksheet_inv['E' + str(writer_pointer)].font = font_important_content
            worksheet_inv['H' + str(writer_pointer)].font = font_important_content
            worksheet_inv['H' + str(writer_pointer)].number_format = '"$"#,##0.00_-'
            writer_pointer = writer_pointer + 1

    for info in model_list:
        if info['2nd Item Number'] == 'MOQ':
            worksheet_inv['E' + str(writer_pointer)] = info['2nd Item Number']
            worksheet_inv['H' + str(writer_pointer)] = int(info['Quantity']) * float(info['Unit Price'])
            total_amount = total_amount + int(info['Quantity']) * float(info['Unit Price'])

            worksheet_inv['E' + str(writer_pointer)].font = font_important_content
            worksheet_inv['H' + str(writer_pointer)].font = font_important_content
            worksheet_inv['H' + str(writer_pointer)].number_format = '"$"#,##0.00_-'
            writer_pointer = writer_pointer + 1

    # 取出表中的各行,第二行到倒数第二行
    for char in range(ord('A'), ord('H')):
        char = chr(char)
        if char == 'B' or char == 'C' or char == 'D':
            continue
        for i in range(model_start_pointer - 1, writer_pointer):
            worksheet_inv[char + str(i)].border = Border(right=side)

    worksheet_inv['C' + str(writer_pointer)] = 'TOTAL'
    worksheet_inv['F' + str(writer_pointer)] = '=SUM(F' + str(model_start_pointer) + ':F' + str(model_end_pointer) + ')'
    worksheet_inv['H' + str(writer_pointer)] = '=SUM(H' + str(model_start_pointer) + ':H' + str(
        writer_pointer - 1) + ')'

    for char in range(ord('A'), ord('I')):
        worksheet_inv[chr(char) + str(writer_pointer)].border = Border(top=side)
    worksheet_inv['H' + str(writer_pointer)].number_format = '"$"#,##0.00_-'

    for cell in worksheet_inv[writer_pointer]:
        cell.font = font_important_content

    for i in range(model_start_pointer, writer_pointer + 1):
        worksheet_inv['B' + str(i)].alignment = Alignment(horizontal='right', vertical='bottom')
        worksheet_inv['D' + str(i)].alignment = Alignment(horizontal='center', vertical='bottom')
        worksheet_inv['F' + str(i)].alignment = Alignment(horizontal='center', vertical='bottom')
        worksheet_inv['G' + str(i)].alignment = Alignment(horizontal='center', vertical='bottom')
        worksheet_inv['H' + str(i)].alignment = Alignment(horizontal='center', vertical='bottom')

    for i in range(model_start_pointer - 1, writer_pointer + 1):
        worksheet_inv.row_dimensions[i].height = 16.5

    # ---表格尾部

    worksheet_inv['E' + str(writer_pointer + 2)] = 'SAY: ' + spell_number(total_amount) + ' U.S. DOLLARS ONLY.'
    worksheet_inv['E' + str(writer_pointer + 4)] = 'THIS SHIPMENT CONTAINS NO SOLID WOOD PACKING MATERIAL'
    worksheet_inv['E' + str(writer_pointer + 6)] = 'Container#:'
    worksheet_inv['E' + str(writer_pointer + 7)] = 'Seal#: '
    worksheet_inv['E' + str(writer_pointer + 8)] = 'BL#'
    for i in range(writer_pointer + 2, writer_pointer + 9):
        worksheet_inv['E' + str(i)].font = font_important_content

    worksheet_inv['G' + str(writer_pointer + 8)] = 'Radio Flyer China Limited.'
    worksheet_inv['G' + str(writer_pointer + 8)].font = Font(name="Arail", size=9, bold=True, italic=True)

    # 设置单元格的边框线条
    border = Border(bottom=side)
    worksheet_inv['G' + str(writer_pointer + 12)].border = border
    worksheet_inv['H' + str(writer_pointer + 12)].border = border
    worksheet_inv['G' + str(writer_pointer + 13)] = 'Carrie Liu'
    worksheet_inv['G' + str(writer_pointer + 14)] = 'Sr. Logistics Specialist'
    worksheet_inv['G' + str(writer_pointer + 13)].font = Font(name="Arail", size=10, bold=True, italic=True)
    worksheet_inv['G' + str(writer_pointer + 14)].font = Font(name="Arail", size=10, bold=True, italic=True)

    manufacturer_list = temp_dict['manufacturer_list']
    for info in manufacturer_list:
        writer_pointer = writer_pointer + 2
        worksheet_inv['A' + str(writer_pointer)] = 'Manufacturer:'
        worksheet_inv['A' + str(writer_pointer)].font = font_important_content

        info_list = info.split(';')
        for i in info_list:
            if i != '':
                writer_pointer = writer_pointer + 1
                worksheet_inv['A' + str(writer_pointer)] = i
                worksheet_inv['A' + str(writer_pointer)].font = Font(name="Arail", size=8)

    writer_pointer = writer_pointer + 2
    worksheet_inv['A' + str(writer_pointer)] = 'Seller:'
    worksheet_inv['A' + str(writer_pointer)].font = font_important_content
    worksheet_inv[
        'A' + str(writer_pointer + 1)] = 'The Radio Flyer Company' + '\n' + '6515 W Grand Ave., Chicago IL 60707, USA'
    worksheet_inv['A' + str(writer_pointer + 1)].font = Font(name="Arail", size=9)
    worksheet_inv['A' + str(writer_pointer + 1)].alignment = Alignment(wrapText=True)
    worksheet_inv.merge_cells("A" + str(writer_pointer + 1) + ':C' + str(writer_pointer + 2))
    for i in range(writer_pointer + 1, writer_pointer + 3):
        worksheet_inv.row_dimensions[i].height = 13

    # ---格式调整

    # 调整列宽
    worksheet_inv.column_dimensions['A'].width = 15
    worksheet_inv.column_dimensions['B'].width = 13.8
    worksheet_inv.column_dimensions['C'].width = 7.6
    worksheet_inv.column_dimensions['D'].width = 2
    worksheet_inv.column_dimensions['E'].width = 40
    worksheet_inv.column_dimensions['F'].width = 14
    worksheet_inv.column_dimensions['G'].width = 14
    worksheet_inv.column_dimensions['H'].width = 20

    # 调整页面
    worksheet_inv.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet_inv.page_setup.fitToWidth = 1
    if item_count <= 7:
        worksheet_inv.page_setup.fitToHeight = 1
    else:
        worksheet_inv.page_setup.fitToHeight = 0

    worksheet_inv.oddHeader.right.text = "ORIGINAL\nCOMMERCIAL INVOICE"  # 文本
    worksheet_inv.oddHeader.right.size = 10  # 字号
    worksheet_inv.oddHeader.right.font = "Arial,Bold"  # 字体
    worksheet_inv.oddFooter.center.text = "Page &[Page] of &[Pages]"  # 文本
    worksheet_inv.oddFooter.center.size = 11  # 字号
    worksheet_inv.oddFooter.center.font = "宋体"  # 字体

    return worksheet_inv


def write_pl_template(worksheet_pl, temp_dict, temp_content_dict):
    # 定义格式（字体、对齐方式）
    font_title = Font(name="Arail", size=10, bold=True, italic=True)
    font_content = Font(name="Arail", size=10)
    align_title_left = Alignment(horizontal='center', vertical='bottom')
    align_content_left = Alignment(vertical='bottom')
    align_title_right = Alignment(horizontal='right', vertical='bottom')
    align_content_right = Alignment(horizontal='left', vertical='bottom')
    title_row_number = 1

    # ---sold to
    # 单个单元格的值写入
    worksheet_pl['A1'] = 'Consign To:'
    worksheet_pl['A1'].font = font_title
    worksheet_pl['A1'].alignment = align_title_left

    # 一列单元格的值写入
    sold_to_address_list = temp_dict['sold_to address info'].split(';')
    writer_pointer = 1
    list_pointer = 0
    while (list_pointer < len(sold_to_address_list)):
        info = sold_to_address_list[list_pointer]
        if info != '':
            worksheet_pl['B' + str(writer_pointer)] = info
            worksheet_pl['B' + str(writer_pointer)].font = font_content
            worksheet_pl['B' + str(writer_pointer)].alignment = align_content_left

            writer_pointer = writer_pointer + 1
            title_row_number += 1
        list_pointer = list_pointer + 1

    # ---ship to
    writer_pointer = writer_pointer + 1
    title_row_number += 1
    # 单个单元格的值写入
    worksheet_pl['A' + str(writer_pointer)] = 'Ship To:'
    worksheet_pl['A' + str(writer_pointer)].font = font_title
    worksheet_pl['A' + str(writer_pointer)].alignment = align_title_left

    # 一列单元格的值写入
    ship_to_address_list = temp_dict['ship_to address info'].split(';')
    list_pointer = 0
    while (list_pointer < len(ship_to_address_list)):
        info = ship_to_address_list[list_pointer]
        if info != '':
            worksheet_pl['B' + str(writer_pointer)] = info
            worksheet_pl['B' + str(writer_pointer)].font = font_content
            worksheet_pl['B' + str(writer_pointer)].alignment = align_content_left

            writer_pointer = writer_pointer + 1
            title_row_number += 1
        list_pointer = list_pointer + 1

    # ---shipped via
    writer_pointer = writer_pointer + 1
    title_row_number += 1
    # 单个单元格的值写入
    worksheet_pl['A' + str(writer_pointer)] = 'Shipped Via:'
    worksheet_pl['A' + str(writer_pointer)].font = font_title
    worksheet_pl['A' + str(writer_pointer)].alignment = align_title_left

    # ---表头右半部分
    worksheet_pl['K1'] = 'Invoice No:'
    worksheet_pl['K2'] = 'Date:'
    worksheet_pl['K3'] = 'Payment Terms:'
    worksheet_pl['K4'] = 'Customer PO#:'
    worksheet_pl['K6'] = 'COUNTRY OF ORIGIN:'
    worksheet_pl['K7'] = 'FOB:'
    worksheet_pl['L1'] = temp_dict['Invoice No']
    worksheet_pl['L2'] = temp_dict['Date'][:-9]
    worksheet_pl['L2'].number_format = 'yyyy-mm-dd'
    worksheet_pl['L3'] = temp_dict['PAYMENT TERM']
    worksheet_pl['L4'] = temp_dict['Customer PO']
    worksheet_pl['L6'] = temp_dict['COUNTRY OF ORIGIN']
    worksheet_pl['L7'] = temp_dict['FOB']

    for i in range(1, 8):
        worksheet_pl['K' + str(i)].font = font_title
        worksheet_pl['L' + str(i)].font = font_content
        worksheet_pl['K' + str(i)].alignment = align_title_right
        worksheet_pl['L' + str(i)].alignment = align_content_right

    for i in range(1, writer_pointer + 1):
        worksheet_pl.row_dimensions[i].height = 15
    worksheet_pl.row_dimensions[writer_pointer + 1].height = 12

    # ---表格部分！！
    font_head_title = Font(name="Arail", size=11, bold=True)
    font_important_content = Font(name="Arail", size=10, bold=True)
    font_trivial_content = Font(name="Arail", size=10)

    # 表格表头
    writer_pointer = writer_pointer + 2
    title_row_number += 2
    worksheet_pl['A' + str(writer_pointer)] = 'Marks & Nos.'
    worksheet_pl['B' + str(writer_pointer)] = 'Description'
    worksheet_pl.merge_cells("B" + str(writer_pointer) + ':E' + str(writer_pointer))
    worksheet_pl['F' + str(writer_pointer)] = 'Quantity in pcs'
    worksheet_pl['G' + str(writer_pointer)] = 'Qty / Carton'
    worksheet_pl['H' + str(writer_pointer)] = 'Cartons'
    worksheet_pl['I' + str(writer_pointer)] = 'Net. WT (KGS)'
    worksheet_pl['J' + str(writer_pointer)] = 'Gross WT. (KGS)'
    worksheet_pl['K' + str(writer_pointer)] = 'CBM'
    worksheet_pl['L' + str(writer_pointer)] = 'G.W (LBS)'
    worksheet_pl['M' + str(writer_pointer)] = 'CFT'
    for cell in worksheet_pl[writer_pointer]:
        cell.font = Font(name="Arail", size=10, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
    worksheet_pl.row_dimensions[writer_pointer].height = 30

    # 设置线条的样式和颜色
    side = Side(style="medium")
    # 设置单元格的边框线条
    worksheet_pl['A' + str(writer_pointer)].border = Border(bottom=side, top=side, right=side)
    worksheet_pl['B' + str(writer_pointer)].border = Border(bottom=side, top=side)
    worksheet_pl['C' + str(writer_pointer)].border = Border(bottom=side, top=side)
    worksheet_pl['D' + str(writer_pointer)].border = Border(bottom=side, top=side)
    worksheet_pl['E' + str(writer_pointer)].border = Border(bottom=side, top=side, right=side)
    worksheet_pl['F' + str(writer_pointer)].border = Border(bottom=side, top=side, right=side)
    worksheet_pl['G' + str(writer_pointer)].border = Border(bottom=side, top=side, right=side)
    worksheet_pl['H' + str(writer_pointer)].border = Border(bottom=side, top=side, right=side)
    worksheet_pl['I' + str(writer_pointer)].border = Border(bottom=side, top=side, right=side)
    worksheet_pl['J' + str(writer_pointer)].border = Border(bottom=side, top=side, right=side)
    worksheet_pl['K' + str(writer_pointer)].border = Border(bottom=side, top=side, right=side)
    worksheet_pl['L' + str(writer_pointer)].border = Border(bottom=side, top=side, right=side)
    worksheet_pl['M' + str(writer_pointer)].border = Border(bottom=side, top=side)

    # 表格内容
    model_list = temp_content_dict
    model_start_pointer = writer_pointer + 2
    title_row_number += 1

    model_index_list = []
    carton_total = 0
    item_count = 0

    for info in model_list:
        if info['2nd Item Number'] != 'MOQ':

            item_count += 1
            if item_count == 8:
                worksheet_pl.print_title_rows = '1:' + str(title_row_number)
                writer_pointer = 69
            else:
                writer_pointer = writer_pointer + 2
            i = writer_pointer

            worksheet_pl['B' + str(writer_pointer)] = 'Item No.'
            worksheet_pl['C' + str(writer_pointer)] = info['2nd Item Number']
            worksheet_pl['D' + str(writer_pointer)] = '-'
            worksheet_pl['E' + str(writer_pointer)] = info['Description']
            worksheet_pl['F' + str(writer_pointer)] = int(info['Quantity'])
            worksheet_pl['G' + str(writer_pointer)] = int(info['Qty/\nCarton'])
            worksheet_pl['H' + str(writer_pointer)] = '=F' + str(writer_pointer) + '/G' + str(writer_pointer)
            worksheet_pl['I' + str(writer_pointer + 1)] = float(info['Net Weight (kg)'])
            worksheet_pl['J' + str(writer_pointer + 1)] = float(info['Gross Weight (kg)'])
            worksheet_pl['K' + str(writer_pointer + 1)] = round(float(info['Cubic\nMeters (per carton)']),4)
            worksheet_pl['L' + str(writer_pointer + 1)] = round(float(info['Gross Weight (lbs)']),2)
            worksheet_pl['M' + str(writer_pointer + 1)] = round(float(info['Cubic\nFeet (per Carton)']),4)
            worksheet_pl['I' + str(writer_pointer)] = '=I' + str(writer_pointer + 1) + '*H' + str(writer_pointer)
            worksheet_pl['J' + str(writer_pointer)] = '=J' + str(writer_pointer + 1) + '*H' + str(writer_pointer)
            worksheet_pl['K' + str(writer_pointer)] = '=K' + str(writer_pointer + 1) + '*H' + str(writer_pointer)
            worksheet_pl['L' + str(writer_pointer)] = '=L' + str(writer_pointer + 1) + '*H' + str(writer_pointer)
            worksheet_pl['M' + str(writer_pointer)] = '=M' + str(writer_pointer + 1) + '*H' + str(writer_pointer)
            model_index_list.append(str(writer_pointer))
            carton_total = carton_total + int(info['Quantity']) / int(info['Qty/\nCarton'])
            # worksheet_pl['G'+str(writer_pointer)].number_format='"$"#,##0.00_-'
            # worksheet_pl['H'+str(writer_pointer)].number_format='"$"#,##0.00_-'

            writer_pointer = writer_pointer + 1
            worksheet_pl['B' + str(writer_pointer)] = 'HTS NO.:'
            worksheet_pl['C' + str(writer_pointer)] = info['For US']
            writer_pointer = writer_pointer + 1
            worksheet_pl['B' + str(writer_pointer)] = 'UPC NO.:'
            worksheet_pl['C' + str(writer_pointer)] = info['12 digits UPC']
            writer_pointer = writer_pointer + 1
            worksheet_pl['B' + str(writer_pointer)] = 'Dimension(cm): '
            worksheet_pl['C' + str(writer_pointer)] = info['Width (L) cm'] + '*' + info['Depth (W) cm'] + '*' + info[
                'Height (H) cm']
            writer_pointer = writer_pointer + 1
            worksheet_pl['B' + str(writer_pointer)] = 'Customer SKU:'
            worksheet_pl['C' + str(writer_pointer)] = info['Cross Reference Item Number']

            for char in range(ord('B'), ord('N')):
                worksheet_pl[chr(char) + str(i)].font = font_important_content
            for char in range(ord('I'), ord('N')):
                worksheet_pl[chr(char) + str(i + 1)].font = font_trivial_content
                if chr(char) == 'K' or chr(char) == 'M':
                    worksheet_pl[chr(char) + str(i + 1)].number_format = '0.0000'
                    worksheet_pl[chr(char) + str(i)].number_format = '0.000'
                else:
                    worksheet_pl[chr(char) + str(i + 1)].number_format = '0.00'
                    worksheet_pl[chr(char) + str(i)].number_format = '0.00'
            for char in range(ord('B'), ord('D')):
                for j in range(i + 1, i + 5):
                    worksheet_pl[chr(char) + str(j)].font = font_trivial_content

    model_end_pointer = writer_pointer
    writer_pointer = writer_pointer + 2

    # 取出表中的各行,第二行到倒数第二行
    for char in range(ord('A'), ord('M')):
        char = chr(char)
        if char == 'B' or char == 'C' or char == 'D':
            continue
        for i in range(model_start_pointer - 1, writer_pointer):
            worksheet_pl[char + str(i)].border = Border(right=side)

    worksheet_pl['C' + str(writer_pointer)] = 'TOTAL'
    worksheet_pl['F' + str(writer_pointer)] = '=SUM(F' + str(model_start_pointer) + ':F' + str(model_end_pointer) + ')'
    worksheet_pl['H' + str(writer_pointer)] = '=SUM(H' + str(model_start_pointer) + ':H' + str(model_end_pointer) + ')'
    worksheet_pl['I' + str(writer_pointer)] = '=I' + '+I'.join(model_index_list)
    worksheet_pl['J' + str(writer_pointer)] = '=J' + '+J'.join(model_index_list)
    worksheet_pl['K' + str(writer_pointer)] = '=K' + '+K'.join(model_index_list)
    worksheet_pl['L' + str(writer_pointer)] = '=L' + '+L'.join(model_index_list)
    worksheet_pl['M' + str(writer_pointer)] = '=M' + '+M'.join(model_index_list)

    for char in range(ord('A'), ord('N')):
        worksheet_pl[chr(char) + str(writer_pointer)].border = Border(top=side)
    for char in range(ord('I'), ord('N')):
        if chr(char) == 'M' or chr(char) == 'K':
            worksheet_pl[chr(char) + str(writer_pointer)].number_format = '#,##0.000_-'
        else:
            worksheet_pl[chr(char) + str(writer_pointer)].number_format = '#,##0.00_-'
    # worksheet_pl['H'+str(writer_pointer)].number_format='"$"#,##0.00_-'

    for cell in worksheet_pl[writer_pointer]:
        cell.font = font_important_content

    for i in range(model_start_pointer, writer_pointer + 1):
        worksheet_pl['B' + str(i)].alignment = Alignment(horizontal='right', vertical='bottom')
        worksheet_pl['D' + str(i)].alignment = Alignment(horizontal='center', vertical='bottom')
        worksheet_pl['F' + str(i)].alignment = Alignment(horizontal='center', vertical='bottom')
        worksheet_pl['G' + str(i)].alignment = Alignment(horizontal='center', vertical='bottom')
        worksheet_pl['H' + str(i)].alignment = Alignment(horizontal='center', vertical='bottom')
        worksheet_pl['I' + str(i)].alignment = Alignment(horizontal='center', vertical='bottom')
        worksheet_pl['J' + str(i)].alignment = Alignment(horizontal='center', vertical='bottom')
        worksheet_pl['K' + str(i)].alignment = Alignment(horizontal='center', vertical='bottom')
        worksheet_pl['L' + str(i)].alignment = Alignment(horizontal='center', vertical='bottom')
        worksheet_pl['M' + str(i)].alignment = Alignment(horizontal='center', vertical='bottom')

    for i in range(model_start_pointer - 1, writer_pointer + 1):
        worksheet_pl.row_dimensions[i].height = 16.5

    # ---表格尾部
    worksheet_pl['E' + str(writer_pointer + 2)] = 'TOTAL: ' + spell_number(round(carton_total, 2)) + '(' + str(
        int(carton_total)) + ') CARTONS ONLY'
    worksheet_pl['E' + str(writer_pointer + 4)] = 'THIS SHIPMENT CONTAINS NO SOLID WOOD PACKING MATERIAL'
    worksheet_pl['E' + str(writer_pointer + 6)] = 'Container#:'
    worksheet_pl['E' + str(writer_pointer + 7)] = 'Seal#: '
    worksheet_pl['E' + str(writer_pointer + 8)] = 'BL#'
    for i in range(writer_pointer + 2, writer_pointer + 9):
        worksheet_pl['E' + str(i)].font = font_important_content

    worksheet_pl['K' + str(writer_pointer + 4)] = 'Radio Flyer China Limited.'
    worksheet_pl['K' + str(writer_pointer + 4)].font = Font(name="Arail", size=9, bold=True, italic=True)

    # 设置单元格的边框线条
    border = Border(bottom=side)
    worksheet_pl['K' + str(writer_pointer + 9)].border = border
    worksheet_pl['L' + str(writer_pointer + 9)].border = border
    worksheet_pl['M' + str(writer_pointer + 9)].border = border
    worksheet_pl['K' + str(writer_pointer + 10)] = 'Carrie Liu'
    worksheet_pl['K' + str(writer_pointer + 11)] = 'Sr. Logistics Specialist'
    worksheet_pl['K' + str(writer_pointer + 10)].font = Font(name="Arail", size=10, bold=True, italic=True)
    worksheet_pl['K' + str(writer_pointer + 11)].font = Font(name="Arail", size=10, bold=True, italic=True)

    manufacturer_list = temp_dict['manufacturer_list']
    for info in manufacturer_list:
        writer_pointer = writer_pointer + 2
        worksheet_pl['A' + str(writer_pointer)] = 'Manufacturer:'
        worksheet_pl['A' + str(writer_pointer)].font = font_important_content

        info_list = info.split(';')
        for i in info_list:
            if i != '':
                writer_pointer = writer_pointer + 1
                worksheet_pl['A' + str(writer_pointer)] = i
                worksheet_pl['A' + str(writer_pointer)].font = Font(name="Arail", size=8)

    writer_pointer = writer_pointer + 2
    worksheet_pl['A' + str(writer_pointer)] = 'Seller:'
    worksheet_pl['A' + str(writer_pointer)].font = font_important_content
    worksheet_pl['A' + str(writer_pointer + 1)] = 'The Radio Flyer Company' + '\n' + '6515 W Grand Ave., Chicago IL 60707, USA'
    worksheet_pl['A' + str(writer_pointer + 1)].font = Font(name="Arail", size=9)
    worksheet_pl['A' + str(writer_pointer + 1)].alignment = Alignment(wrapText=True)
    worksheet_pl.merge_cells("A" + str(writer_pointer + 1) + ':C' + str(writer_pointer + 2))
    for i in range(writer_pointer + 1, writer_pointer + 3):
        worksheet_pl.row_dimensions[i].height = 14

    # ---格式调整

    # 调整列宽
    worksheet_pl.column_dimensions['A'].width = 12
    worksheet_pl.column_dimensions['B'].width = 14
    worksheet_pl.column_dimensions['C'].width = 7.6
    worksheet_pl.column_dimensions['D'].width = 2
    worksheet_pl.column_dimensions['E'].width = 41
    worksheet_pl.column_dimensions['F'].width = 8.5
    worksheet_pl.column_dimensions['G'].width = 7
    worksheet_pl.column_dimensions['H'].width = 8.5
    worksheet_pl.column_dimensions['I'].width = 10
    worksheet_pl.column_dimensions['J'].width = 10
    worksheet_pl.column_dimensions['K'].width = 10
    worksheet_pl.column_dimensions['L'].width = 10
    worksheet_pl.column_dimensions['M'].width = 10

    # 调整页面
    worksheet_pl.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet_pl.page_setup.fitToWidth = 1
    if item_count <= 7:
        worksheet_pl.page_setup.fitToHeight = 1
    else:
        worksheet_pl.page_setup.fitToHeight = 0

    worksheet_pl.oddHeader.right.text = "ORIGINAL\nPACKING SLIP"  # 文本
    worksheet_pl.oddHeader.right.size = 10  # 字号
    worksheet_pl.oddHeader.right.font = "Arial,Bold"  # 字体
    worksheet_pl.oddFooter.center.text = "Page &[Page] of &[Pages]"  # 文本
    worksheet_pl.oddFooter.center.size = 11  # 字号
    worksheet_pl.oddFooter.center.font = "宋体"  # 字体
    return worksheet_pl


def write_excel(table_head_dict, model_dict):
    for index in range(len(table_head_dict)):
        file_name = list(table_head_dict.items())[index][0]
        temp_dict = list(table_head_dict.items())[index][1]
        temp_content_dict = list(model_dict.items())[index][1]

        # 创建一个 ExcelWriter 对象，指定 engine='openpyxl' 参数
        writer = pd.ExcelWriter(folder_path + '\\documents\\' + str(file_name) + '.xlsx', engine='openpyxl')

        # 将 DataFrame 写入 Excel 文件
        pd.DataFrame().to_excel(writer, sheet_name='Inv', index=False)
        pd.DataFrame().to_excel(writer, sheet_name='PL', index=False)
        workbook = writer.book

        worksheet_inv = writer.sheets['Inv']
        worksheet_inv = write_inv_template(worksheet_inv, temp_dict, temp_content_dict)
        worksheet_pl = writer.sheets['PL']
        worksheet_pl = write_pl_template(worksheet_pl, temp_dict, temp_content_dict)
        writer.close()


if __name__ == '__main__':
    root = tk.Tk()
    root.wm_attributes('-topmost', 1)
    root.withdraw()
    messagebox.showinfo("提示", "程序启动中！")

    folder_path = os.getcwd()
    table_data = data_extraction(file_dir=folder_path, data_file_name='\\original data.xlsx')
    table_head_dict, model_dict = data_integration(data=table_data)
    write_excel(table_head_dict, model_dict)
    messagebox.showinfo("提示", "Completed.")
